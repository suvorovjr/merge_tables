import openpyxl
import pandas as pd
import os
import csv
from datetime import datetime
from pathlib import Path


class ToCSVClient:

    def __init__(self, market_file):
        self.market_file = market_file

    def clear_excel_files(self):
        workbook = openpyxl.load_workbook(self.market_file)
        sheet = workbook.active
        ToCSVClient.share_cells(sheet)
        ToCSVClient.delete_unwanted_row(sheet)
        new_path = ToCSVClient.get_to_csv(sheet, self.market_file)
        return new_path

    @staticmethod
    def share_cells(sheet):
        merged_cells_ranges = list(sheet.merged_cells)
        for merged_range in merged_cells_ranges:
            sheet.unmerge_cells(str(merged_range))
        sheet.delete_rows(6)

    @staticmethod
    def is_colored(cell):
        if cell.fill.fgColor.rgb is not None and cell.fill.fgColor.rgb != "00000000" and cell.fill.fgColor.rgb != "FFFFFFFF":
            return True
        return False

    @staticmethod
    def get_to_csv(sheet, excel_path):
        csv_path = os.path.splitext(excel_path)[0] + '.csv'
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as file:
            c = csv.writer(file, delimiter=';')
            for row in sheet.iter_rows(min_row=5, values_only=True):
                c.writerow(row)
        return csv_path

    @staticmethod
    def delete_unwanted_row(sheet):
        for row in sheet.iter_rows(min_row=5, max_row=5):
            for cell in row:
                if not ToCSVClient.is_colored(cell):
                    sheet.delete_cols(cell.column, 1)


class MergeFiles:

    def __init__(self, san_file, market_file):
        self.san_file = san_file
        self.market_file = market_file

    def merge_files(self):
        csv_client = ToCSVClient(self.market_file)
        market_csv_path = csv_client.clear_excel_files()
        san_df = pd.read_csv(self.san_file, delimiter=';')
        market_df = pd.read_csv(market_csv_path, delimiter=';')
        merge_df = pd.merge(san_df, market_df, left_on='SKU на нашем сайте', right_on='SKU', how='inner')
        save_dir = Path(__file__).parent.parent / 'media/merge_files'
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        merge_filename = f'{save_dir}/merge_{datetime.now()}.xlsx'
        merge_df.to_excel(merge_filename, index=False)
        return merge_filename
